"""
数据处理和Excel生成工具
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from config import DATA_DIR, EXCEL_TEMPLATE

class DataProcessor:
    """数据处理类"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.wechat_data = []
        self.baidu_search_data = []
        self.baidu_info_data = []
        
    def process_baidu_data(self, raw_data):
        """处理百度指数数据"""
        try:
            self.logger.info("开始处理百度指数数据")
            
            # 这里需要根据实际收集到的数据结构进行处理
            # 示例数据结构
            if 'search_data' in raw_data:
                search_data = raw_data['search_data']
                self.baidu_search_data = self._parse_baidu_index_data(search_data)
            
            if 'info_data' in raw_data:
                info_data = raw_data['info_data']
                self.baidu_info_data = self._parse_baidu_index_data(info_data)
            
            self.logger.info("百度指数数据处理完成")
            return True
            
        except Exception as e:
            self.logger.error(f"处理百度指数数据失败: {str(e)}")
            return False
    
    def process_wechat_data(self, raw_data):
        """处理微信指数数据"""
        try:
            self.logger.info("开始处理微信指数数据")
            
            if raw_data.get('method') == 'web' and 'data' in raw_data:
                self.wechat_data = self._parse_wechat_index_data(raw_data['data'])
            elif raw_data.get('method') == 'manual':
                # 手动收集的数据，需要用户手动输入
                self.logger.info("微信指数数据需要手动输入")
                self.wechat_data = []
            
            self.logger.info("微信指数数据处理完成")
            return True
            
        except Exception as e:
            self.logger.error(f"处理微信指数数据失败: {str(e)}")
            return False
    
    def _parse_baidu_index_data(self, raw_data):
        """解析百度指数数据"""
        try:
            parsed_data = []
            
            # 这里根据实际数据结构进行解析
            # 示例：假设返回的是字典格式 {date: {keyword: value}}
            if isinstance(raw_data, dict):
                for date_str, keyword_data in raw_data.items():
                    row = {'日期': date_str}
                    for keyword, value in keyword_data.items():
                        row[f'{keyword}'] = value
                    parsed_data.append(row)
            
            return parsed_data
            
        except Exception as e:
            self.logger.error(f"解析百度指数数据失败: {str(e)}")
            return []
    
    def _parse_wechat_index_data(self, raw_data):
        """解析微信指数数据"""
        try:
            parsed_data = []
            
            # 这里根据实际数据结构进行解析
            for item in raw_data:
                if 'keyword' in item and 'data' in item:
                    parsed_data.append(item)
            
            return parsed_data
            
        except Exception as e:
            self.logger.error(f"解析微信指数数据失败: {str(e)}")
            return []
    
    def calculate_weekly_averages(self, data, keyword):
        """计算每周平均值"""
        try:
            if not data:
                return {}
            
            # 将数据转换为DataFrame
            df = pd.DataFrame(data)
            
            # 确保日期列为datetime格式
            df['日期'] = pd.to_datetime(df['日期'])
            
            # 按周分组计算平均值
            df['周数'] = df['日期'].dt.isocalendar().week
            
            weekly_avg = df.groupby('周数')[keyword].mean().to_dict()
            
            return weekly_avg
            
        except Exception as e:
            self.logger.error(f"计算每周平均值失败: {str(e)}")
            return {}
    
    def generate_excel_report(self, output_path):
        """生成Excel报告"""
        try:
            self.logger.info("开始生成Excel报告")
            
            # 创建Excel写入器
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # 1. 生成微信指数趋势表
                self._generate_wechat_sheet(writer)
                
                # 2. 生成百度指数搜索表
                self._generate_baidu_search_sheet(writer)
                
                # 3. 生成百度指数资讯表
                self._generate_baidu_info_sheet(writer)
                
                # 4. 生成汇总表
                self._generate_summary_sheet(writer)
            
            # 应用样式
            self._apply_excel_styles(output_path)
            
            self.logger.info(f"Excel报告已生成: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"生成Excel报告失败: {str(e)}")
            return False
    
    def _generate_wechat_sheet(self, writer):
        """生成微信指数工作表"""
        try:
            if not self.wechat_data:
                # 创建空模板
                df = self._create_empty_template('微信指数趋势')
                df.to_excel(writer, sheet_name='微信指数趋势', index=False)
                return
            
            # 处理微信指数数据
            df = pd.DataFrame(self.wechat_data)
            
            # 添加周几列
            df['日期'] = pd.to_datetime(df['日期'])
            df['星期'] = df['日期'].dt.day_name()
            
            # 计算每周平均值
            for keyword in ['上海电信', '上海移动', '上海联通']:
                if keyword in df.columns:
                    weekly_avg = self.calculate_weekly_averages(df.to_dict('records'), keyword)
                    df[f'{keyword}_每周平均'] = df.apply(
                        lambda row: weekly_avg.get(row['日期'].dt.isocalendar().week[0], ''),
                        axis=1
                    )
            
            # 重新排列列
            columns = ['日期', '星期', '上海电信', '上海电信_每周平均', 
                      '上海移动', '上海移动_每周平均', 
                      '上海联通', '上海联通_每周平均']
            
            # 确保所有列都存在
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[columns]
            df.to_excel(writer, sheet_name='微信指数趋势', index=False)
            
        except Exception as e:
            self.logger.error(f"生成微信指数工作表失败: {str(e)}")
    
    def _generate_baidu_search_sheet(self, writer):
        """生成百度指数搜索工作表"""
        try:
            if not self.baidu_search_data:
                # 创建空模板
                df = self._create_empty_template('百度指数搜索')
                df.to_excel(writer, sheet_name='百度指数搜索', index=False)
                return
            
            # 处理百度指数搜索数据
            df = pd.DataFrame(self.baidu_search_data)
            
            # 添加周几列
            df['日期'] = pd.to_datetime(df['日期'])
            df['星期'] = df['日期'].dt.day_name()
            
            # 计算每周平均值
            for keyword in ['上海电信', '上海移动', '上海联通']:
                if keyword in df.columns:
                    weekly_avg = self.calculate_weekly_averages(df.to_dict('records'), keyword)
                    df[f'{keyword}_每周平均'] = df.apply(
                        lambda row: weekly_avg.get(row['日期'].dt.isocalendar().week[0], ''),
                        axis=1
                    )
            
            # 重新排列列
            columns = ['日期', '星期', '上海电信', '上海电信_每周平均', 
                      '上海移动', '上海移动_每周平均', 
                      '上海联通', '上海联通_每周平均']
            
            # 确保所有列都存在
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[columns]
            df.to_excel(writer, sheet_name='百度指数搜索', index=False)
            
        except Exception as e:
            self.logger.error(f"生成百度指数搜索工作表失败: {str(e)}")
    
    def _generate_baidu_info_sheet(self, writer):
        """生成百度指数资讯工作表"""
        try:
            if not self.baidu_info_data:
                # 创建空模板
                df = self._create_empty_template('百度指数资讯')
                df.to_excel(writer, sheet_name='百度指数资讯', index=False)
                return
            
            # 处理百度指数资讯数据
            df = pd.DataFrame(self.baidu_info_data)
            
            # 添加周几列
            df['日期'] = pd.to_datetime(df['日期'])
            df['星期'] = df['日期'].dt.day_name()
            
            # 计算每周平均值
            for keyword in ['上海电信', '上海移动', '上海联通']:
                if keyword in df.columns:
                    weekly_avg = self.calculate_weekly_averages(df.to_dict('records'), keyword)
                    df[f'{keyword}_每周平均'] = df.apply(
                        lambda row: weekly_avg.get(row['日期'].dt.isocalendar().week[0], ''),
                        axis=1
                    )
            
            # 重新排列列
            columns = ['日期', '星期', '上海电信', '上海电信_每周平均', 
                      '上海移动', '上海移动_每周平均', 
                      '上海联通', '上海联通_每周平均']
            
            # 确保所有列都存在
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            
            df = df[columns]
            df.to_excel(writer, sheet_name='百度指数资讯', index=False)
            
        except Exception as e:
            self.logger.error(f"生成百度指数资讯工作表失败: {str(e)}")
    
    def _generate_summary_sheet(self, writer):
        """生成汇总工作表"""
        try:
            summary_data = []
            
            # 汇总各运营商的指数数据
            operators = ['上海电信', '上海移动', '上海联通']
            platforms = ['微信指数', '百度指数搜索', '百度指数资讯']
            
            for operator in operators:
                row = {'运营商': operator}
                
                # 微信指数平均值
                if self.wechat_data:
                    df_wechat = pd.DataFrame(self.wechat_data)
                    if operator in df_wechat.columns:
                        row['微信指数平均'] = df_wechat[operator].mean()
                
                # 百度指数搜索平均值
                if self.baidu_search_data:
                    df_search = pd.DataFrame(self.baidu_search_data)
                    if operator in df_search.columns:
                        row['百度指数搜索平均'] = df_search[operator].mean()
                
                # 百度指数资讯平均值
                if self.baidu_info_data:
                    df_info = pd.DataFrame(self.baidu_info_data)
                    if operator in df_info.columns:
                        row['百度指数资讯平均'] = df_info[operator].mean()
                
                summary_data.append(row)
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='数据汇总', index=False)
            
        except Exception as e:
            self.logger.error(f"生成汇总工作表失败: {str(e)}")
    
    def _create_empty_template(self, sheet_type):
        """创建空模板"""
        try:
            if '微信' in sheet_type:
                columns = ['日期', '星期', '上海电信', '每周指数平均值', 
                          '上海移动', '每周指数平均值', 
                          '上海联通', '每周指数平均值', '三家指数趋势对比']
            elif '搜索' in sheet_type:
                columns = ['日期', '星期', '上海电信每日搜索指数', '每周指数平均值',
                          '上海移动每日搜索指数', '每周指数平均值',
                          '上海联通每日搜索指数', '每周指数平均值', '三家每日搜索指数趋势对比']
            else:  # 资讯
                columns = ['日期', '星期', '上海电信每日资讯指数', '每周指数平均值',
                          '上海移动每日资讯指数', '每周指数平均值',
                          '上海联通每日资讯指数', '每周指数平均值', '三家每日资讯指数趋势对比']
            
            # 创建7天的空数据
            dates = []
            start_date = datetime.now() - timedelta(days=7)
            for i in range(7):
                date = start_date + timedelta(days=i)
                dates.append({
                    '日期': date.strftime('%Y-%m-%d'),
                    '星期': date.strftime('%A')
                })
            
            df = pd.DataFrame(dates)
            
            # 添加其他空列
            for col in columns:
                if col not in df.columns:
                    df[col] = ''
            
            return df[columns]
            
        except Exception as e:
            self.logger.error(f"创建空模板失败: {str(e)}")
            return pd.DataFrame()
    
    def _apply_excel_styles(self, filepath):
        """应用Excel样式"""
        try:
            wb = load_workbook(filepath)
            
            # 定义样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            peak_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            peak_font = Font(color="FFFFFF", bold=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为每个工作表应用样式
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置标题行样式
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # 设置数据行样式
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 如果是数值列，检查是否为峰值
                        if isinstance(cell.value, (int, float)) and cell.value:
                            # 这里可以添加峰值检测逻辑
                            pass
            
            wb.save(filepath)
            self.logger.info(f"Excel样式应用完成: {filepath}")
            
        except Exception as e:
            self.logger.error(f"应用Excel样式失败: {str(e)}")

def main():
    """主函数"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    processor = DataProcessor()
    
    # 示例：生成空模板
    output_path = f"/mnt/okcomputer/output/index_collector/data/运营商指数报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    if processor.generate_excel_report(output_path):
        print(f"Excel报告生成成功: {output_path}")
    else:
        print("Excel报告生成失败")

if __name__ == '__main__':
    main()